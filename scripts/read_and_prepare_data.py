import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO


def extraer_tabla_historico(df: pd.DataFrame) -> pd.DataFrame:
    """
    Extrae y limpia la tabla de resumen por categorías desde un DataFrame.
    Incluye 'Total', marca si es total y si la fila es roja en el Excel.
    """

    end_row = df[df.iloc[:, 0].astype(str).str.contains("Total", case=False, na=False)].index.min()
    tabla = df.iloc[1:end_row + 1, [0, 1, 4, 5, 6, 18, 19, 21]].copy()
    tabla.columns = ["Categoría", "Recibidas", "Atendidas_num", "Atendidas_%", "Duracion",
                     "Desborde_cantidad", "Desborde_tiempo", "Abandonadas"]

    # Marcar fila 'Total'
    tabla["EsTotal"] = tabla["Categoría"].astype(str).str.contains("Total", case=False, na=False)

    # Marcar si es roja según openpyxl (usamos índice original)


    # Filtrar solo las que tienen % válido
    tabla = tabla[tabla["Atendidas_%"].astype(str).str.contains("%", na=False)]

    # Limpiar y convertir valores
    tabla["Atendidas_%"] = (
        tabla["Atendidas_%"].astype(str)
        .str.replace(",", ".")
        .str.replace(" %", "", regex=False)
        .astype(float)
    )

    return tabla.reset_index(drop=True)

def extraer_tabla_categorias(path_excel: str) -> pd.DataFrame:

    df = pd.read_csv(path_excel, encoding='ISO-8859-1', delimiter=';')

    
    print("Hojas disponibles")
    tabla = df.iloc[:, [0, 1, 4, 5, 6, 18, 19, 21]].copy()
    tabla.columns = ["Categoría", "Recibidas", "Atendidas_num", "Atendidas_%", "Duracion",
                     "Desborde_cantidad", "Desborde_tiempo", "Abandonadas"]
    # Limpiar y convertir porcentaje
    tabla = tabla[tabla["Atendidas_%"].astype(str).str.contains("%", na=False)]

    tabla["Atendidas_%"] = (
        tabla["Atendidas_%"]
        .astype(str)
        .str.replace(",", ".")
        .str.replace(" %", "", regex=False)
        .astype(float)
    )

    return tabla.reset_index(drop=True), df
    

import pandas as pd
from datetime import datetime

def extraer_tabla_dias(path_excel: str) -> pd.DataFrame:
    """
    Extrae y limpia la tabla por día desde la primera hoja de un Excel.
    Añade la columna 'Día' (nombre del día en español) como primera columna, excepto en la fila final (total).
    """

    # Diccionario para traducir los días de la semana
    dias_semana = {
        "Monday": "Lunes",
        "Tuesday": "Martes",
        "Wednesday": "Miércoles",
        "Thursday": "Jueves",
        "Friday": "Viernes",
        "Saturday": "Sábado",
        "Sunday": "Domingo",
    }

    # Leer la primera hoja del Excel
    df = pd.read_csv(path_excel, encoding='ISO-8859-1', delimiter=';')

    # Seleccionar columnas relevantes
    tabla = df.iloc[:, [0, 1, 4, 5]].copy()
    tabla.columns = ["Categoría", "Recibidas", "Atendidas_num", "Atendidas_%"]

    # Limpiar y convertir porcentaje
    tabla = tabla[tabla["Atendidas_%"].astype(str).str.contains("%", na=False)]
    tabla["Atendidas_%"] = (
        tabla["Atendidas_%"]
        .astype(str)
        .str.replace(",", ".")
        .str.replace(" %", "", regex=False)
        .astype(float)
    )

    # Añadir columna Día (excepto para la última fila)
    dias = []
    for i in range(len(tabla)):
        if i < len(tabla) - 1:
            try:
                dia = dias_semana[datetime.strptime(tabla.loc[i, "Categoría"], "%d-%m-%Y").strftime("%A")]
            except Exception:
                dia = ""
        else:
            dia = ""
        dias.append(dia)

    tabla.insert(0, "Día", dias)
    df.insert(0, "Día", dias)

    return tabla.reset_index(drop=True), df



def extraer_tabla_franjas(path_excel: str) -> pd.DataFrame:
    """
    Extrae y limpia la tabla por franja horaria desde un DataFrame.
    Detecta si la fila está escrita en rojo, incluye columna EsTotal y limpia %.
    """
    df = pd.read_csv(path_excel, encoding='ISO-8859-1', delimiter=';')


    # Extraer tabla
    tabla = df.iloc[:, [0, 1, 4, 5]].copy()
    tabla.columns = ["Categoría", "Recibidas", "Atendidas_num", "Atendidas_%"]


    # Limpiar porcentaje
    tabla = tabla[tabla["Atendidas_%"].astype(str).str.contains("%", na=False)]
    tabla["Atendidas_%"] = (
        tabla["Atendidas_%"]
        .astype(str)
        .str.replace(",", ".")
        .str.replace(" %", "", regex=False)
        .astype(float)
    )

    return tabla.reset_index(drop=True), df

import pandas as pd
import streamlit as st

def leer_tablas(path_paquete: str, meses: list[str]) -> dict:
    """
    Lee las tablas de categorías, días y franjas horarias desde un archivo Excel.
    """
    st.write("📥 Paso 1: Accediendo a paths individuales")
    path_excel_historico = path_paquete[3]
    path_excel_categorias = path_paquete[0]
    path_excel_dias = path_paquete[1]
    path_excel_franjas = path_paquete[2]

    st.write("✅ Paths extraídos correctamente")

    tablas_mes = {}
    for m in meses[:-1]:  # Excluimos el mes actual
        st.write(f"📊 Leyendo hoja histórica del mes: {m}")
        try:
            df = pd.read_excel(path_excel_historico, sheet_name=m, header=None)
            tablas_mes[m] = extraer_tabla_historico(df)
            st.write(f"✅ Hoja '{m}' leída correctamente con shape {df.shape}")
        except Exception as e:
            st.error(f"❌ Error al leer hoja '{m}': {e}")
            raise

    st.write("📌 Paso 2: leyendo tablas individuales...")

    try:
        categorias_filtrado, df_categorias = extraer_tabla_categorias(path_excel_categorias)
        st.write("✅ Tabla de categorías leída:", df_categorias.shape)
    except Exception as e:
        st.error(f"❌ Error en categorías: {e}")
        raise

    try:
        dias_filtrado, df_dias = extraer_tabla_dias(path_excel_dias)
        st.write("✅ Tabla de días leída:", df_dias.shape)
    except Exception as e:
        st.error(f"❌ Error en días: {e}")
        raise

    try:
        franjas_filtrado, df_franjas = extraer_tabla_franjas(path_excel_franjas)
        st.write("✅ Tabla de franjas leída:", df_franjas.shape)
    except Exception as e:
        st.error(f"❌ Error en franjas: {e}")
        raise

    tablas = {
        "categorias": categorias_filtrado,
        "dias": dias_filtrado,
        "franjas": franjas_filtrado
    }

    st.write("📌 Paso 3: Generando Excel con resumen")

    try:
        excel = append_dfs_to_excel_path(
            path_excel_historico,
            df_categorias,
            df_dias,
            df_franjas,
            sheet_name=meses[-1]  # Último mes
        )
        st.write("✅ Excel actualizado generado correctamente")
    except Exception as e:
        st.error(f"❌ Error al generar Excel actualizado: {e}")
        raise

    st.write("✅ Todas las tablas leídas y procesadas correctamente")

    return tablas, tablas_mes, excel



def append_dfs_to_excel_path(excel_path: str, df1, df2, df3, sheet_name="Resumen") -> BytesIO:
    # Load workbook from path
    wb = load_workbook(filename=excel_path)

    # If sheet exists, delete it
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    # Create new sheet
    ws = wb.create_sheet(title=sheet_name)

    # Write DataFrames with 3-row spacing
    start_row = 1
    for df in [df1, df2, df3]:
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        start_row += len(df) + 4  # 3 blank rows + 1 header

    # Save to BytesIO to return through Streamlit
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
if __name__ == "__main__":
    # Ejemplo de uso
    MESES = [ "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
                       "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE" ]
    path_excel = "data/INFORME DE LLAMADAS AUTOCLIMA ABRIL 2025.xlsx"
    nombre_hoja = "ABRIL"
    meses_hasta_ahora = MESES[:MESES.index(nombre_hoja)]

    leer_tablas(path_excel, meses_hasta_ahora)
